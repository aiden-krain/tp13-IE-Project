import geojson
from geojson import Feature, FeatureCollection, Polygon
import numpy as np
import folium
import openpyxl
import pandas as pd

class ReadExcle(object):
    def __init__(self, filname, sheet_name):

        self.filename = filname
        self.sheet_name = sheet_name
    def open(self):
        self.wb = openpyxl.load_workbook(self.filename)
        self.sh = self.wb[self.sheet_name]

    def read_data(self):
        result = {}
        self.open()
        max_row = self.sh.max_row
        max_colmn = self.sh.max_column
        datas = []
        for i in range(1, max_row + 1):
            date = []
            for j in range(1, max_colmn + 1):
                date1 = self.sh.cell(i, j).value
                date.append(date1)
            datas.append(date)
        for ij in datas:
            result[ij[0].upper()] = ij[1]
        return result

excel = ReadExcle(r"data.xlsx", "Sheet3")
result = excel.read_data()

print(result)

with open('111.json', 'rb') as f:
    districts = geojson.load(f)
features = []
color = []
vlist = {}
for idx, geometry in enumerate(districts['features']):
    circle = geometry['geometry']['coordinates'][0][0]
    vlist[str(idx)] =str(geometry['properties']['vic_loca_2'])
    color.append([str(idx), np.random.randint(100)])
    polygon = Polygon([circle])
    features.append(Feature(
        id=str(idx),
        geometry=polygon
    ))

feature_collection = FeatureCollection(features)
#print(vlist)
#print(color)
result_color = []
for i in color:
    if i[0] in vlist:
        z = vlist[i[0]]
        if z in result:
            i[1] = result[z]
            result_color.append(i)
        else:
            i[1]=0
            result_color.append(i)
    else:
        i[1] = 0
        result_color.append(i)
color = result_color
#print(color)
map = folium.Map(
    location=[-37.81,144],

    tiles=('https://api.tiles.mapbox.com/v4/mapbox.streets/{z}/{x}/{y}.png'
           + '?access_token=pk.eyJ1IjoibHVrYXNtYXJ0aW5lbGxpIiwiYSI6ImNpem8'
           + '5dmhwazAyajIyd284dGxhN2VxYnYifQ.HQCmyhEXZUTz3S98FMrVAQ'),
    attr='Melbourne|TP13'
)
folium.Choropleth(
    geo_data=feature_collection,
    name='choropleth',
    data=pd.DataFrame(color, columns=['idx','color']),
    columns=['idx','color'],
    key_on='feature.id',
    fill_color='BuPu',
    fill_opacity=0.5,
    line_opacity=0.2,
    legend_name='Rank'
).add_to(map)


map.save('templates/map.html')
